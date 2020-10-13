import openpyxl
from django.http import HttpResponse
import os
import mimetypes


def download(excel_file_name, data):
    # создаем новый excel-файл
    wb = openpyxl.Workbook()
    # добавляем новый лист
    wb.create_sheet(title='main', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['main']
    sheet.column_dimensions['A'].width = 60
    for row_index, row in enumerate(data):
        for col_index, value in enumerate(row):
            cell = sheet.cell(row=(row_index + 1), column=(col_index+1))
            cell.value = value
    wb.save(excel_file_name)
    fp = open(excel_file_name, "rb")
    response = HttpResponse(fp.read())
    fp.close()
    file_type = mimetypes.guess_type(excel_file_name)
    if file_type is None:
        file_type = 'application/octet-stream'
    response['Content-Type'] = file_type
    response['Content-Length'] = str(os.stat(excel_file_name).st_size)
    response['Content-Disposition'] = "attachment; filename={}".format(excel_file_name)
    os.remove(excel_file_name)
    return response
